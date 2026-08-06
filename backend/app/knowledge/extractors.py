from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import TYPE_CHECKING, Any
from xml.etree import ElementTree

if TYPE_CHECKING:
    from app.knowledge.ocr import TesseractOcrEngine


TEXT_EXTENSIONS = {
    ".c",
    ".cfg",
    ".conf",
    ".cpp",
    ".cs",
    ".css",
    ".csv",
    ".go",
    ".h",
    ".html",
    ".ini",
    ".java",
    ".js",
    ".json",
    ".jsx",
    ".kt",
    ".log",
    ".md",
    ".php",
    ".properties",
    ".ps1",
    ".py",
    ".rb",
    ".rs",
    ".rst",
    ".sh",
    ".sql",
    ".toml",
    ".ts",
    ".tsx",
    ".txt",
    ".xml",
    ".yaml",
    ".yml",
}
OFFICE_EXTENSIONS = {".docx", ".pptx", ".xlsx", ".odt"}
SUPPORTED_EXTENSIONS = TEXT_EXTENSIONS | OFFICE_EXTENSIONS | {".pdf"}


@dataclass(frozen=True)
class ExtractedText:
    text: str
    encoding: str
    extractor: str
    extractor_version: str | None = None
    processor_variant: str | None = None


class SpreadsheetExtractionLimitError(ValueError):
    def __init__(self, reason_code: str, message: str) -> None:
        super().__init__(message)
        self.reason_code = reason_code


def extract_text(path: Path) -> str:
    return extract_text_with_metadata(path).text


def extract_text_with_metadata(
    path: Path,
    *,
    max_spreadsheet_cells: int = 250_000,
    max_output_characters: int = 10_000_000,
    ocr_engine: TesseractOcrEngine | None = None,
) -> ExtractedText:
    suffix = path.suffix.lower()
    if suffix in TEXT_EXTENSIONS:
        raw = path.read_bytes()
        encoding = detect_text_encoding(raw)
        return ExtractedText(raw.decode(encoding), encoding, "text")
    if suffix == ".xlsx":
        return _extract_xlsx(
            path,
            max_cells=max_spreadsheet_cells,
            max_characters=max_output_characters,
        )
    if suffix in OFFICE_EXTENSIONS:
        return ExtractedText(
            _extract_zipped_xml(path),
            "office-xml",
            "office-xml",
        )
    if suffix == ".pdf":
        text = _extract_pdf(path)
        if text.strip() or ocr_engine is None:
            return ExtractedText(text, "pdf-text", "pypdf")
        result = ocr_engine.extract_pdf(path)
        return ExtractedText(
            result.text,
            "pdf-ocr",
            result.engine,
            result.engine_version,
            f"pdf_ocr_v1:{result.languages}:{result.pages}",
        )
    raise ValueError(f"Unsupported knowledge file type: {suffix}")


def detect_text_encoding(raw: bytes) -> str:
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if raw.startswith(b"\xff\xfe"):
        return "utf-16-le"
    if raw.startswith(b"\xfe\xff"):
        return "utf-16-be"
    try:
        raw.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        pass
    for encoding in ("cp932", "shift_jis"):
        try:
            raw.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError(
        "supported-enterprise-text",
        raw,
        0,
        min(1, len(raw)),
        "expected UTF-8, UTF-16, CP932, or Shift-JIS",
    )


def _extract_zipped_xml(path: Path) -> str:
    parts: list[str] = []
    with zipfile.ZipFile(path) as archive:
        for name in sorted(archive.namelist()):
            if not name.endswith(".xml"):
                continue
            if not (
                name.startswith("word/")
                or name.startswith("ppt/slides/")
                or name.startswith("xl/sharedStrings")
                or name.startswith("xl/worksheets/")
                or name == "content.xml"
            ):
                continue
            root = ElementTree.fromstring(archive.read(name))
            text = " ".join(
                value.strip()
                for value in root.itertext()
                if value and value.strip()
            )
            if text:
                parts.append(text)
    return "\n".join(parts)


def _xlsx_value(value: Any) -> str:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    text = str(value)
    return (
        text.replace("\\", "\\\\")
        .replace("\r", "\\r")
        .replace("\n", "\\n")
        .replace("\t", "\\t")
    )


def _extract_xlsx(
    path: Path,
    *,
    max_cells: int,
    max_characters: int,
) -> ExtractedText:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "XLSX indexing requires the openpyxl package"
        ) from exc

    formula_workbook = None
    value_workbook = None
    lines: list[str] = [f"[workbook] name={_xlsx_value(path.name)}"]
    character_count = len(lines[0])
    populated_cells = 0

    def append_line(line: str) -> None:
        nonlocal character_count
        next_count = character_count + len(line) + 1
        if next_count > max_characters:
            raise SpreadsheetExtractionLimitError(
                "spreadsheet_text_limit_exceeded",
                "Spreadsheet extracted text exceeds the configured limit",
            )
        lines.append(line)
        character_count = next_count

    try:
        formula_workbook = openpyxl.load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        value_workbook = openpyxl.load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        for sheet_index, formula_sheet in enumerate(
            formula_workbook.worksheets,
            start=1,
        ):
            value_sheet = value_workbook[formula_sheet.title]
            append_line(
                "[sheet] "
                f"index={sheet_index} "
                f"name={_xlsx_value(formula_sheet.title)} "
                f"state={formula_sheet.sheet_state}"
            )
            formula_rows = formula_sheet.iter_rows()
            value_rows = value_sheet.iter_rows()
            for formula_row, value_row in zip(
                formula_rows,
                value_rows,
                strict=False,
            ):
                for formula_cell, value_cell in zip(
                    formula_row,
                    value_row,
                    strict=False,
                ):
                    if formula_cell.value is None:
                        continue
                    populated_cells += 1
                    if populated_cells > max_cells:
                        raise SpreadsheetExtractionLimitError(
                            "spreadsheet_cell_limit_exceeded",
                            "Spreadsheet populated cell count exceeds "
                            "the configured limit",
                        )
                    coordinate = formula_cell.coordinate
                    if formula_cell.data_type == "f":
                        formula = _xlsx_value(formula_cell.value)
                        line = f"{coordinate}\tformula={formula}"
                        if value_cell.value is not None:
                            line += (
                                "\tcached_value="
                                + _xlsx_value(value_cell.value)
                            )
                    else:
                        line = (
                            f"{coordinate}\tvalue="
                            + _xlsx_value(formula_cell.value)
                        )
                    append_line(line)
    finally:
        if formula_workbook is not None:
            formula_workbook.close()
        if value_workbook is not None:
            value_workbook.close()

    return ExtractedText(
        "\n".join(lines),
        "xlsx-semantic",
        "openpyxl",
        openpyxl.__version__,
        "xlsx_semantic_v1",
    )


def _extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
        from pypdf.errors import PdfReadError
    except ImportError as exc:
        raise RuntimeError(
            "PDF indexing requires the pypdf package"
        ) from exc
    try:
        reader = PdfReader(str(path))
        parts = []
        for index, page in enumerate(reader.pages, start=1):
            page_text = page.extract_text() or ""
            if page_text.strip():
                parts.append(f"[page] index={index}\n{page_text}")
        return "\n".join(parts)
    except PdfReadError as exc:
        raise ValueError(f"PDF cannot be read: {exc}") from exc


def normalize_text(text: str) -> str:
    text = text.replace("\x00", "")
    text = re.sub(r"\r\n?", "\n", text)
    return "\n".join(line.rstrip() for line in text.splitlines()).strip()
